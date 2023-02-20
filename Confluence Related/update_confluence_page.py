""" 

The following should be modified for script to work efficiently:
- PUT_EXCEL_SHEET_LOCATION_HERE
- PUT_CREDENTIALS_LOCATION_HERE
  - Use the "CREDENTIALS_FILE" template in neighoring file (https://github.com/CFMCGEE/ChatGBT_Stuffs/blob/main/Confluence%20Related/CF_CREDENTIALS.txt)
- MASTER_FOLDER

Install the following non-built in py libraries (https://pypi.org/):
- pip install openpyxl
- pip install pywin32
- pip install pytz
- pip install atlassian-python-api
- pip install bs4

 """

import re
import time
import json
import openpyxl
import win32com.client as win32
from pytz import timezone
from atlassian import Confluence
from datetime import datetime, date
from bs4 import BeautifulSoup as BS

# Load the Excel file
path_to_credentials = 'PUT_CREDENTIALS_LOCATION_HERE' # 
path_to_excel_sheet = 'PUT_EXCEL_SHEET_LOCATION_HERE'
wb = openpyxl.load_workbook(filename=path_to_excel_sheet)
ws = wb.active

def get_confluence_credentials(credentials_path):

    # Get the path to the credentials file
    file_path = credentials_path 

    # Read the file contents
    with open(file_path, 'r') as f:
        file_contents = f.read()

    # Split the file contents by newlines and strip any leading/trailing whitespace
    lines = [line.strip() for line in file_contents.split('\n')]

    # Parse the lines to get the API token and username
    api_token = None
    username = None

    for line in lines:
        if line.startswith('API_TOKEN='):
            api_token = line.split('=', 1)[1]
        elif line.startswith('USERNAME='):
            username = line.split('=', 1)[1]

    return api_token, username

def refresh_excel_sheet_and_compare(confluence_instance, page_id, page_content, excel_sheet, excel_sheet_path):

    # Remove HTML tags from the page content
    cleaned_content = re.sub('<[^<]+?>', '', page_content).replace("&amp;", "&")

    # Opens Excel file alternatively
    excel = win32.Dispatch("Excel.Application")
    wb = excel.Workbooks.Open(excel_sheet_path)
    ws = wb.ActiveSheet

    # Refresh sheet
    wb.RefreshAll()

    # Waits 5 seconds before checking for new data
    print(f'Waiting 5 seconds after refreshing Excel sheet checking...')
    time.sleep(5)

    # Closes Excel file
    wb.Close()
    excel.Quit()

    # Waits 5 seconds before re-opening excel sheet
    print(f'\nWaiting 5 seconds before re-opening excel sheet...')
    time.sleep(5)

    # Open the Excel sheet
    workbook = excel_sheet
    sheet = workbook.active

    files_being_added_to_confluence_page = []

    # Compare Confluence content with Excel sheet content
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] in cleaned_content:
            continue # <--- Skips
        else:
            files_being_added_to_confluence_page.append(row[0])
            print(f"{row[0]} is not found on the Confluence page.")

    return str(f"The following files need to be added to the Confluence Page: {files_being_added_to_confluence_page}")

def get_est_time():

    """Get the current Eastern Standard Time (EST)"""
    est = timezone('US/Eastern')
    now = datetime.now(est)

    return now.strftime("%I:%M %p").lstrip('0') + " EST"

def get_confluence_instance(url, username, token):

    """Set up a Confluence instance with API token"""
    return Confluence(url=url, username=username, password=token)

def get_page_name(confluence_instance, page_id):

    """Get the name of a Confluence page by ID"""
    return confluence_instance.get_page_by_id(page_id=page_id, expand='title')['title']

def get_page_content(confluence_instance, page_id):

    """Get the content of a Confluence page by ID"""
    return confluence_instance.get_page_by_id(page_id=page_id, expand='body.storage')['body']['storage']['value']

def get_latest_version_comment(confluence_instance, page_id):

    """Get the latest version comment of a Confluence page by ID"""
    history = confluence_instance.history(page_id)
    return history['lastUpdated']['message']

def get_next_version_number(version_comment):

    """Get the next version number based on the latest version comment"""
    match = re.search(r'\d+', version_comment)
    if match:
        return int(match.group()) + 1
    else:
        return 1

def get_user_input(page_name, version_number):

    """Prompt the user for input and confirm it"""
    while True:
        user_input = input(f"Please enter what updates that have been made to the '{page_name}' Page: ")
        confirm_input = input(f"You entered '{user_input} [{version_number}]'. Is this correct? (y/n): ")
        if confirm_input.lower() == "y":
            return user_input, version_number
        elif confirm_input.lower() == "n":
            print(f"Please try again and confirm your updates that have been made to the '{page_name}' Page.")
        else:
            print("Invalid input. Please try again.")

def generate_table_html(confluence_instance, page_id, column_headers, folder, files, date_created, date_of_last_modification, local_location, time_of_upload, date_of_upload):
    
    """Generate HTML code for a table"""
    new_table_html = '<table>' \
                    '<tr>' + ''.join([f'<th style="width:auto">{header}</th>' for header in column_headers]) + '</tr>' \
                    '<tr>' + ''.join([f'<td>{value}</td>' for value in [folder, files, date_created, date_of_last_modification, local_location, time_of_upload, date_of_upload]]) + '</tr>' \
                    '</table>'

    return new_table_html

def update_confluence_page(confluence_instance, page_id, page_title, new_table_content, version_comment):

    """Update a Confluence page with new content"""
    return confluence_instance.update_page(page_id, page_title, new_table_content, version_comment=version_comment, minor_edit=True)

def parse_worksheet(ws):

    """Parses a worksheet and returns a dictionary of data grouped by folder path."""
    divided_folders = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        path = re.sub(r"^.*MASTER_FOLDER", "", "{v}".format(v=row[5])).replace("\\", "")
        if path not in divided_folders:
            divided_folders[path] = []
        divided_folders[path].append({
            'Folder Path': path,
            'File Name': row[0],
            'File Type': row[1],
            'Date Created': row[3].strftime('%m/%d/%Y'),
            'Date of Last Modification': row[4].strftime('%m/%d/%Y'),
            'Location (Path)': row[5]
        })

    return divided_folders

def main():

    """Main function"""

    url = 'https://cfpracticesite.atlassian.net'
    page_id = 65538
    space_key = 'MFS'
    column_headers = ['Folder', 'Files', 'Date Created', 'Date of Last Modification', 'Local Location (Path)', 'Time of Upload', 'Date of Upload']

    api_token, username = get_confluence_credentials(path_to_credentials)
    confluence_instance = get_confluence_instance(url, username, api_token)
    page_name = get_page_name(confluence_instance, page_id)
    page_content = get_page_content(confluence_instance, page_id)
    latest_version_comment = get_latest_version_comment(confluence_instance, page_id)
    next_version_number = get_next_version_number(latest_version_comment)
    user_input, version_number = get_user_input(page_name, next_version_number)

    # Refresh Excel file content
    refresh_excel_sheet_and_compare(confluence_instance, page_id, page_content, wb, path_to_excel_sheet)

    # Column headers as HTML 
    column_headers_html = '<table><tr><th style="width:auto">Folder</th><th style="width:auto">Files</th><th style="width:auto">Date Created</th><th style="width:auto">Date of Last Modification</th><th style="width:auto">Local Location (Path)</th><th style="width:auto">Time of Upload</th><th style="width:auto">Date of Upload</th></tr>'

    # Folders parser
    folders = parse_worksheet(ws)
    content_to_add = ""

    for folder_name in folders:

        # Input values
        folder = []
        files = ""
        date_created = []
        date_of_last_modification = []
        local_location = []
        time_of_upload = get_est_time().lstrip('0')
        date_of_upload = date.today().strftime('%m/%d/%Y')

        # Loop through folders and get the folder name, file name, and file type
        for fi in folders[folder_name]:
            folder.append(fi['Folder Path'])
            files += "{0}, ".format(fi['File Name'])
            date_created.append(fi['Date Created'])
            date_of_last_modification.append(fi['Date of Last Modification'])
            local_location.append(fi['Location (Path)'])

        # Creates HTML table
        table_content = generate_table_html(
            confluence_instance, 
            page_id, 
            column_headers, 
            folder[0], 
            files.rsplit(',', 1)[0],
            date_created[0], 
            date_of_last_modification[0], 
            local_location[0], 
            time_of_upload, 
            date_of_upload
            )

        # Adds reformatted table content to string     
        content_to_add += table_content.replace(column_headers_html, "").replace("</table>", "").replace("&", "&amp;")
   
    # New HTML page
    new_table_content = column_headers_html + content_to_add + "</table>"

    # Updates Confluence Page
    update_confluence_page(confluence_instance, page_id, page_name, new_table_content, version_number)

    # Prettifies the HTML table
    soup = BS(new_table_content, 'html.parser')
    new_table_content_reformatted = soup.prettify()
    
    # Prints HTML table content
    print("New Table Content: \n {con}".format(con=new_table_content_reformatted))
    
    # Print confirmation that the page has been updated and print out new table content
    wb.close()
    print(f"The page '{page_name}' has been updated successfully & excel has been closed!")

main()

""" 

Purpose: This script ought to be executed daily to verify and identify newly added folders and items, with the aim of updating the Confluence Page.

Logic & Functionalty still yet to be implemented:

- Removing files if they are removed from Confluence Page
- Rather than updating entire table, only add rows to existing table
- Checks for sub-folders in MASTER_FOLDER
- Change user from file path to MASTER_FOLDER's account username
- Timer set to refresh after period of time (automation later on?)

"""
